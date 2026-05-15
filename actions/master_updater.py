import csv
import os
from loguru import logger
from openpyxl import load_workbook
from resources.config import Config


def _read_csv(file_path: str) -> list:
    """Return all rows from a UTF-8-BOM-safe CSV as a list of lists."""
    with open(file_path, encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


def _coerce(value: str):
    """Convert a string to int or float where possible, else keep as str."""
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    return value


def _clear_range(ws, cell_range: str) -> None:
    """Blank every cell in the given A1-notation range."""
    for row in ws[cell_range]:
        for cell in row:
            cell.value = None


def _write_rows(ws, start_row: int, start_col: int, rows: list) -> None:
    """Write a list-of-lists into the sheet starting at (start_row, start_col)."""
    for i, row in enumerate(rows):
        for j, value in enumerate(row):
            ws.cell(row=start_row + i, column=start_col + j, value=_coerce(value))


def update_master_from_csvs(config: Config, date_str: str) -> str:
    """Update the master template using CSVs for the given date and save a dated copy."""
    dated_output_dir = os.path.join(config.output_dir, date_str)
    collections_csv = os.path.join(dated_output_dir, f"{date_str}_collections.csv")
    registrations_csv = os.path.join(dated_output_dir, f"{date_str}_registrations_state_wise.csv")
    claim_csv = os.path.join(dated_output_dir, f"{date_str}_claim_status.csv")

    if not os.path.exists(dated_output_dir):
        raise FileNotFoundError(f"Output folder not found: {dated_output_dir}")

    missing = [
        path for path in [collections_csv, registrations_csv, claim_csv] if not os.path.exists(path)
    ]
    if missing:
        missing_list = "\n".join(missing)
        raise FileNotFoundError(f"Missing CSV files:\n{missing_list}")

    if not os.path.exists(config.master_template):
        raise FileNotFoundError(f"Master template not found: {config.master_template}")

    wb = load_workbook(config.master_template)
    ws = wb["Sheet1"]

    #Collections: Skip header row; write to V1:Y36
    collections_rows = _read_csv(collections_csv)[1:]
    _clear_range(ws, "V1:Y36")
    _write_rows(ws, 1, 22, collections_rows)
    logger.info(f"Collections: {len(collections_rows)} rows written to V1:Y36")

    #Registrations: Skip title row, header row, and last summary row; write to K2:T15
    all_reg_rows = _read_csv(registrations_csv)
    reg_data_rows = all_reg_rows[2:-1]
    _clear_range(ws, "K2:T15")
    _write_rows(ws, 2, 11, reg_data_rows)
    logger.info(f"Registrations: {len(reg_data_rows)} rows written to K2:T15")

    #Claim status: Write all rows to V37:AA47
    claim_rows = _read_csv(claim_csv)
    _clear_range(ws, "V37:AA47")
    _write_rows(ws, 37, 22, claim_rows)
    logger.info(f"Claim status: {len(claim_rows)} rows written to V37:AA47")

    output_file = os.path.join(dated_output_dir, f"{date_str}_master.xlsx")
    wb.save(output_file)
    logger.success(f"Master updated: {output_file}")
    return output_file
